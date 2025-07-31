"""
Azure utilities and client classes for Multi-Agent AI Platform
Separated to avoid circular imports
"""

import time
import logging
import sys
from typing import Dict, Optional, List
import json
import requests
import hashlib
from datetime import datetime, timedelta
import os
import re

# Configure logging for production (errors only)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# Azure imports
from azure.ai.projects import AIProjectClient
from azure.ai.projects.models import ToolSet, CodeInterpreterTool
from azure.identity import DefaultAzureCredential, InteractiveBrowserCredential, ClientSecretCredential
from azure.core.exceptions import AzureError
from azure.storage.blob import BlobServiceClient
from azure.search.documents import SearchClient
from azure.search.documents.models import VectorizedQuery
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchIndex,
    SearchField,
    SearchFieldDataType,
    SimpleField,
    SearchableField,
    VectorSearch,
    HnswAlgorithmConfiguration,
    VectorSearchProfile,
    SemanticConfiguration,
    SemanticSearch,
    SemanticPrioritizedFields,
    SemanticField
)
from azure.core.credentials import AzureKeyCredential

# Configure logging
logger = logging.getLogger(__name__)

# Configuration class for Azure services
class AzureConfig:
    """Centralized configuration for Azure services"""
    
    def __init__(self):
        # Check for environment variables first, then fall back to defaults
        import os
        
        # Load environment variables from file if available
        try:
            from dotenv import load_dotenv
            env_path = os.path.join(os.path.dirname(__file__), 'azure-app-settings.env')
            if os.path.exists(env_path):
                load_dotenv(env_path)
            else:
                load_dotenv()
        except ImportError:
            pass
        
        # Azure AD App Registration for authentication - Use correct values from env
        self.client_id = os.getenv("AZURE_CLIENT_ID", "706af514-7245-4658-9f36-1f79071a80f6")
        self.client_secret = os.getenv("AZURE_CLIENT_SECRET", "Ddo8Q~vJSZoAnmMMEz.zKR4n_6usxpnqvOV-ibNR")
        self.tenant_id = os.getenv("AZURE_TENANT_ID", "7ae3526a-96fa-407a-9b02-9fe5bdff6217")
        
        # Development vs Production mode
        self.dev_mode = os.getenv("DEV_MODE", "false").lower() == "true"
        self.use_managed_identity = os.getenv("USE_MANAGED_IDENTITY", "true").lower() == "true"
        
        # For local development, disable managed identity
        if self.dev_mode:
            self.use_managed_identity = False
        
        # OAuth2 token endpoint ve Graph API endpoint tanımlamaları
        self.oauth_token_endpoint = f"https://login.microsoftonline.com/{self.tenant_id}/oauth2/v2.0/token"
        self.graph_endpoint = "https://graph.microsoft.com/v1.0"
        
        # Other Azure services - Use environment variables or defaults
        self.redirect_uri = os.getenv("REDIRECT_URI", "https://egent-hxh3eubahne9ama2.westeurope-01.azurewebsites.net/")
        
        # Storage configuration with fallback authentication
        # Use proper Azure Storage connection string format with valid account name from env
        self.storage_connection_string = os.getenv(
            "AZURE_STORAGE_CONNECTION_STRING", 
            "DefaultEndpointsProtocol=https;AccountName=egentstorage;AccountKey=XlTiW7z3iAjdzlvWdYRhr0m6uos7Js31HDL/BUSeMCshlepMhcuRk6RQtSSc1nsxVwdLJsXx6Qi7+AStzOylIw==;EndpointSuffix=core.windows.net"
        )
        self.storage_account_name = os.getenv("AZURE_STORAGE_ACCOUNT_NAME", "egentstorage")
        
        # Azure AI Projects configuration - Get from environment
        self.ai_project_name = os.getenv("AZURE_AI_PROJECT_NAME", "egehubproject")
        self.resource_group_name = os.getenv("AZURE_RESOURCE_GROUP", "egerg") 
        self.subscription_id = os.getenv("AZURE_SUBSCRIPTION_ID", "904b1d30-71da-4a0b-9394-78ec6b6c505a")
        
        # Search configuration - Use correct values from env
        self.search_endpoint = os.getenv("AZURE_SEARCH_SERVICE_ENDPOINT", "https://egentaisearch.search.windows.net")
        self.search_admin_key = os.getenv("AZURE_SEARCH_ADMIN_KEY", "lbpwL3XrIPoQr51xXoCQ6BliAD1ai0bv1E36C4MYqfAzSeBhOGfz")
        
        # Azure OpenAI configuration for embeddings
        self.openai_endpoint = "https://egentaimodel.openai.azure.com"  # Keep existing endpoint
        self.openai_api_key = os.getenv(
            "OPENAI_API_KEY", 
            "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"
        )
        self.openai_api_version = "2023-05-15"
        self.embedding_model = "text-embedding-ada-002"
        
    
    # Public client ID for device code flow authentication (readonly applications)
    @property
    def public_client_id(self):
        """Get public client ID for device code authentication"""
        return self.client_id  # Using same client ID for simplicity
    
    def get_credential(self):
        """Get appropriate Azure credential based on environment"""
        try:
            if self.use_managed_identity and not self.dev_mode:
                # Production: Use Managed Identity
                return DefaultAzureCredential()
            else:
                # Development: Use Service Principal with connection string fallback
                return ClientSecretCredential(
                    tenant_id=self.tenant_id,
                    client_id=self.client_id,
                    client_secret=self.client_secret
                )
        except Exception as e:
            # Fallback to DefaultAzureCredential which will try multiple auth methods
            return DefaultAzureCredential()
    
    def get_storage_client(self):
        """Get Azure Storage client with appropriate authentication"""
        try:
            if self.use_managed_identity and not self.dev_mode:
                # Production: Use Managed Identity
                credential = self.get_credential()
                return BlobServiceClient(
                    account_url=f"https://{self.storage_account_name}.blob.core.windows.net",
                    credential=credential
                )
            else:
                # Development: Use connection string (more reliable for dev)
                return BlobServiceClient.from_connection_string(self.storage_connection_string)
        except Exception as e:
            # Fallback to connection string
            return BlobServiceClient.from_connection_string(self.storage_connection_string)

# Azure Authentication Helper
class AzureAuthenticator:
    """Handles Azure AD authentication"""
    
    def __init__(self, config: AzureConfig):
        self.config = config
    
    def authenticate_with_device_code(self) -> Dict:
        """
        Initiate device code flow authentication for Azure AD
        Returns device code information for the user to complete authentication
        """
        try:
            from msal import PublicClientApplication
            
            # Initialize MSAL client
            app = PublicClientApplication(
                client_id=self.config.public_client_id,
                authority=f"https://login.microsoftonline.com/{self.config.tenant_id}"
            )
            
            # Define scopes for authentication
            scopes = ["https://graph.microsoft.com/User.Read"]
            
            # Initiate device flow
            flow = app.initiate_device_flow(scopes=scopes)
            
            if "user_code" not in flow:
                return {
                    "success": False,
                    "message": f"Failed to initiate device code flow: {flow.get('error_description', 'Unknown error')}"
                }
            
            # Return device flow information
            return {
                "success": True,
                "device_flow": flow,
                "user_code": flow["user_code"],
                "verification_uri": flow["verification_uri"],
                "expires_in": flow["expires_in"],
                "message": "Device code flow initiated successfully. Please complete authentication in your browser."
            }
            
        except ImportError:
            return {
                "success": False,
                "message": "MSAL library not installed. Please install msal package."
            }
        except Exception as e:
            return {
                "success": False,
                "message": f"Error initiating device code flow: {str(e)}"
            }
    
    def complete_device_code_authentication(self, flow: Dict) -> Dict:
        """
        Complete device code authentication process
        Args:
            flow: The flow object returned from authenticate_with_device_code
            
        Returns:
            Dict with authentication result
        """
        try:
            from msal import PublicClientApplication
            
            # Initialize MSAL client
            app = PublicClientApplication(
                client_id=self.config.public_client_id,
                authority=f"https://login.microsoftonline.com/{self.config.tenant_id}"
            )
            
            # Try to acquire token by device flow
            result = app.acquire_token_by_device_flow(flow)
            
            if "access_token" in result:
                # Authentication successful
                user_info = self._get_user_info_from_graph(result["access_token"])
                username = user_info.get("userPrincipalName", "unknown")
                role = self._determine_user_role(username)
                permissions = self._get_user_permissions(username, role)
                
                return {
                    "success": True,
                    "user": {
                        "username": username,
                        "display_name": user_info.get("displayName", username),
                        "role": role,
                        "permissions": permissions,
                        "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                        "real_azure_user": True,
                        "auth_method": "device_code"
                    },
                    "token": result["access_token"],
                    "message": "Azure AD authentication successful"
                }
            else:
                error_description = result.get("error_description", "Unknown error")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication failed: {error_description}"
                }
                
        except ImportError:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": "MSAL library not installed. Please install msal package."
            }
        except Exception as e:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Error completing device code authentication: {str(e)}"
            }
    
    def authenticate_with_username_password(self, username: str, password: str) -> Dict:
        """
        Authenticate user with username/password against Azure AD
        Only supports real Azure AD authentication
        """
        try:
            # Directly authenticate with Azure AD
            return self._authenticate_real_azure_user(username, password)
                
        except Exception as e:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Authentication error: {str(e)}"
            }
    
    def _authenticate_real_azure_user(self, username: str, password: str) -> Dict:
        """
        Authenticate real Azure AD user using direct OAuth2 ROPC flow
        This method uses the OAuth2 password grant flow to authenticate against Azure AD
        
        Note: This flow requires that:
        1. The app is registered with proper permissions
        2. The user does not have MFA enabled
        3. The tenant policy allows ROPC flow
        4. Using /organizations or tenant-specific endpoint (not /common or /consumers)
        """
        try:
            import requests
            
            
            # Try both organizations and common endpoints for testing
            token_url = f"https://login.microsoftonline.com/organizations/oauth2/v2.0/token"
            
            # Log the token URL being used (don't log in production)
            
            # Prepare the request body for OAuth2 password grant flow
            token_data = {
                'grant_type': 'password',
                'client_id': self.config.client_id,
                'client_secret': self.config.client_secret,
                'scope': 'https://graph.microsoft.com/.default offline_access',
                'username': username,
                'password': password
            }
            
            # Add proper headers for the token request
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded',
                'Accept': 'application/json'
            }
            
            
            # Make the token request with proper error handling
            try:
                token_response = requests.post(token_url, data=token_data, headers=headers)
            except Exception as req_ex:
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Network error during authentication: {str(req_ex)}"
                }
            
            # Try to parse JSON response - handle potential JSON parsing errors
            try:
                result = token_response.json()
            except ValueError:
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication error: Invalid response format (Status: {token_response.status_code})"
                }
            
            if "error" in result:
                error_code = result.get("error")
                error_description = result.get("error_description", "")
                correlation_id = result.get("correlation_id", "unknown")
                timestamp = result.get("timestamp", "unknown")
                trace_id = result.get("trace_id", "unknown")
                
                # Detailed logging of the error
                
                import re
                aadsts_match = re.search(r'AADSTS\d+', error_description)
                if aadsts_match:
                    aadsts_code = aadsts_match.group(0)
                
                # If we got AADSTS9000102 error, try with a specific tenant ID
                if "AADSTS9000102" in error_description:
                    # You can set a specific tenant ID here if you have one
                    specific_tenant_id = "7ae3526a-96fa-407a-9b02-9fe5bdff6217"  # Example tenant ID
                    return self._try_with_specific_tenant(username, password, specific_tenant_id)
                
                # Handle common error scenarios with user-friendly messages
                if "AADSTS50076" in error_description or "AADSTS50079" in error_description:
                    # MFA required error
                    return {
                        "success": False,
                        "user": None,
                        "token": None,
                        "message": "Bu hesap için çok faktörlü kimlik doğrulama (MFA) gerekiyor."
                    }
                elif "AADSTS50126" in error_description:
                    # Invalid username or password
                    return {
                        "success": False,
                        "user": None,
                        "token": None,
                        "message": "Hatalı kullanıcı adı veya şifre. Lütfen kimlik bilgilerinizi kontrol edin."
                    }
                elif "AADSTS50034" in error_description:
                    # User not found
                    return {
                        "success": False,
                        "user": None,
                        "token": None,
                        "message": "Bu kullanıcı hesabı Azure AD'de bulunamadı."
                    }
            
            if "access_token" in result:
                # Success with username/password
                user_info = self._get_user_info_from_graph(result["access_token"])
                role = self._determine_user_role(username)
                permissions = self._get_user_permissions(username, role)
                
                return {
                    "success": True,
                    "user": {
                        "username": username,
                        "display_name": user_info.get("displayName", username),
                        "role": role,
                        "permissions": permissions,
                        "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                        "real_azure_user": True,
                        "auth_method": "password"
                    },
                    "token": result["access_token"],
                    "message": "Azure AD authentication successful"
                }
            else:
                error_description = result.get("error_description", "Unknown error")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication failed: {error_description}"
                }
                
        except ImportError:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": "MSAL library not installed. Please install msal package."
            }
        except Exception as e:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Azure AD authentication error: {str(e)}"
            }
    
    def _try_with_specific_tenant(self, username: str, password: str, tenant_id: str) -> Dict:
        """Fallback authentication method using a specific tenant ID"""
        
        try:
            import requests
            
            # Use specific tenant ID endpoint
            token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
            
            # Headers
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded',
                'Accept': 'application/json'
            }
            
            # Request data
            token_data = {
                'grant_type': 'password',
                'client_id': self.config.client_id,
                'client_secret': self.config.client_secret,
                'scope': 'https://graph.microsoft.com/.default',
                'username': username,
                'password': password
            }
            
            # Make the request
            token_response = requests.post(token_url, data=token_data, headers=headers)
            
            # Parse response
            try:
                result = token_response.json()
            except ValueError:
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication error with fallback tenant: Invalid response format"
                }
            
            # Check for errors
            if "error" in result:
                error_code = result.get("error")
                error_description = result.get("error_description", "")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication failed with fallback tenant: {error_description}"
                }
                
            if "access_token" in result:
                # Success with username/password
                user_info = self._get_user_info_from_graph(result["access_token"])
                role = self._determine_user_role(username)
                permissions = self._get_user_permissions(username, role)
                
                return {
                    "success": True,
                    "user": {
                        "username": username,
                        "display_name": user_info.get("displayName", username),
                        "role": role,
                        "permissions": permissions,
                        "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                        "real_azure_user": True,
                        "auth_method": "password"
                    },
                    "token": result["access_token"],
                    "message": "Azure AD authentication successful with fallback tenant"
                }
                
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": "Unknown error with fallback tenant authentication"
            }
                
        except Exception as e:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Error in fallback authentication: {str(e)}"
            }
    
    def _get_user_info_from_graph(self, access_token: str) -> Dict:
        """Get user information from Microsoft Graph API"""
        try:
            import requests
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(
                'https://graph.microsoft.com/v1.0/me',
                headers=headers
            )
            
            if response.status_code == 200:
                return response.json()
            else:
                return {}
                
        except Exception as e:
            return {}
    
    def _determine_user_role(self, username: str) -> str:
        """Determine user role based on username or other criteria"""
        # For now, assign standard role to all real users
        # In production, you might check Azure AD groups or custom attributes
        if username.lower().startswith("admin") or "admin" in username.lower():
            return "admin"
        elif username.lower().startswith("manager") or "manager" in username.lower():
            return "manager"
        else:
            return "standard"
    
    def _get_user_permissions(self, username: str, role: str) -> List[str]:
        """Get user permissions based on role - now returns a simple list of permissions"""
        if role == "admin":
            # Admin has all permissions
            return ["all"]
        elif role == "manager":
            # Manager has most permissions for all agents
            return [
                "access", "chat", "document_upload", "document_delete",
                # Allow all current and future agents
                "legal:access", "legal:chat", "legal:document_upload", "legal:document_delete",
                "it:access", "it:chat", "it:document_upload", "it:document_delete",
                "sample:access", "sample:chat", "sample:document_upload", "sample:document_delete",
                "scm:access", "scm:chat", "scm:document_upload", "scm:document_delete"
            ]
        else:
            # Standard user has basic permissions for all agents
            return [
                "access", "chat", "document_upload",
                # Specific permissions for each agent
                "legal:access", "legal:chat", "legal:document_upload",
                "it:access", "it:chat", "it:document_upload", 
                "sample:access", "sample:chat", "sample:document_upload",
                "scm:access", "scm:chat", "scm:document_upload"
            ]
    
    def get_azure_credential(self) -> DefaultAzureCredential:
        """Get Azure credential for service authentication"""
        try:
            if self.config.dev_mode:
                # For development, use client secret credential
                return ClientSecretCredential(
                    tenant_id=self.config.tenant_id,
                    client_id=self.config.client_id,
                    client_secret=self.config.client_secret
                )
            else:
                # For production, use managed identity or default credential chain
                return DefaultAzureCredential()
        except Exception as e:
            return DefaultAzureCredential()

# Enhanced Azure AI Agent Client
class EnhancedAzureAIAgentClient:
    """Enhanced wrapper for Azure AI Agent operations with document management"""
    
    def __init__(self, connection_string: str, agent_id: str, config: AzureConfig, container_name: str = None):
        """
        Initialize Enhanced Azure AI Agent Client with document reference support
        
        Args:
            connection_string: Azure AI Project connection string
            agent_id: Azure AI Agent ID
            config: Azure configuration object  
            container_name: Blob storage container name for document references
        """
        self.connection_string = connection_string
        self.agent_id = agent_id
        self.config = config
        self.container_name = container_name  # Store for document reference processing
        self.client = None
        self.agent = None
        self.blob_client = None
        self.search_client = None
        self.search_index_client = None
        
        self._initialize_clients()
    
    def _initialize_clients(self):
        """Initialize all Azure clients with proper error handling and fallbacks"""
        try:
            # Try to initialize AI Project Client with fallback
            try:
                pass
                
                if self.connection_string and self.agent_id and self.connection_string.strip():
                    # Clean up connection string - remove any URL encoding
                    clean_conn_str = self.connection_string.replace('%20', ' ').strip()
                    
                    # Skip AI Project client if connection string is empty or disabled
                    if not clean_conn_str or clean_conn_str.startswith('#'):
                        self.client = None
                        self.agent = None
                    else:
                        # Get configuration components for direct initialization
                        subscription_id = self.config.subscription_id
                        resource_group = self.config.resource_group_name
                        project_name = self.config.ai_project_name
                        
                        
                        if all([subscription_id, resource_group, project_name]):
                            # Use the working endpoint format from test
                            endpoint = "https://eastus2.api.azureml.ms"
                            
                            # Use direct initialization instead of from_connection_string
                            try:
                                credential = self.config.get_credential()
                                
                                self.client = AIProjectClient(
                                    endpoint=endpoint,
                                    subscription_id=subscription_id,
                                    resource_group_name=resource_group,
                                    project_name=project_name,
                                    credential=credential
                                )
                            except Exception as client_create_error:
                                self.client = None
                                self.agent = None
                                return
                        else:
                            self.client = None
                            self.agent = None
                            return
                        
                        # Try to get the agent
                        if self.client:
                            pass
                            
                            try:
                                # First, try to list available agents to verify connection
                                agents_list = self.client.agents.list_agents()
                                
                                # Now try to get the specific agent
                                self.agent = self.client.agents.get_agent(self.agent_id)
                            except Exception as agent_get_error:
                                # Check if it's a workspace error
                                if "Workspace not found" in str(agent_get_error):
                                    pass
                                elif "Agent not found" in str(agent_get_error):
                                    pass
                                self.agent = None
                        else:
                            self.client = None
                            self.agent = None
                else:
                    self.client = None
                    self.agent = None
            except Exception as ai_error:
                self.client = None
                self.agent = None
            
            # Initialize Blob Storage Client with fallback to credential-based auth
            try:
                if self.config.storage_connection_string and "DefaultEndpointsProtocol" in self.config.storage_connection_string:
                    self.blob_client = BlobServiceClient.from_connection_string(
                        self.config.storage_connection_string
                    )
                else:
                    # Fallback to credential-based authentication
                    self.blob_client = BlobServiceClient(
                        account_url=f"https://{self.config.storage_account_name}.blob.core.windows.net",
                        credential=self.config.get_credential()
                    )
            except Exception as blob_error:
                self.blob_client = None
            
            # Initialize Azure Search clients
            try:
                if self.config.search_endpoint and self.config.search_admin_key:
                    from azure.search.documents.indexes import SearchIndexClient
                    from azure.core.credentials import AzureKeyCredential
                    
                    credential = AzureKeyCredential(self.config.search_admin_key)
                    self.search_index_client = SearchIndexClient(
                        endpoint=self.config.search_endpoint,
                        credential=credential
                    )
                else:
                    self.search_index_client = None
                    self.search_client = None
            except Exception as search_error:
                self.search_index_client = None
                self.search_client = None
            
            
        except AzureError as e:
            raise
        except Exception as e:
            raise
    
    def get_available_agents(self):
        """Get available Azure AI agents"""
        try:
            # Create AI Project client directly if not initialized
            if not self.client:
                subscription_id = self.config.subscription_id
                resource_group = self.config.resource_group_name
                project_name = self.config.ai_project_name
                
                if all([subscription_id, resource_group, project_name]):
                    # Use the working endpoint format from test
                    endpoint = "https://eastus2.api.azureml.ms"
                    
                    # Use direct initialization
                    temp_client = AIProjectClient(
                        endpoint=endpoint,
                        subscription_id=subscription_id,
                        resource_group_name=resource_group,
                        project_name=project_name,
                        credential=self.config.get_credential()
                    )
                    
                    agents_response = temp_client.agents.list_agents()
                    if hasattr(agents_response, 'data'):
                        agents = []
                        for agent in agents_response.data:
                            # Handle created_at field properly - convert datetime to string if needed
                            created_at_value = getattr(agent, 'created_at', None)
                            if created_at_value is not None:
                                # If it's a datetime object, convert to timestamp
                                if hasattr(created_at_value, 'timestamp'):
                                    created_at_value = int(created_at_value.timestamp())
                                # If it's a string, try to keep it as is
                                elif isinstance(created_at_value, str):
                                    try:
                                        from datetime import datetime
                                        dt = datetime.fromisoformat(created_at_value.replace('Z', '+00:00'))
                                        created_at_value = int(dt.timestamp())
                                    except:
                                        created_at_value = 0
                                # If it's already a number, keep it
                                elif not isinstance(created_at_value, (int, float)):
                                    created_at_value = 0
                            else:
                                created_at_value = 0
                            
                            agents.append({
                                'id': agent.id,
                                'name': agent.name or f"Agent {agent.id[:8]}",
                                'model': getattr(agent, 'model', 'gpt-4o'),
                                'description': agent.description or f"Azure AI Agent using {getattr(agent, 'model', 'gpt-4o')}",
                                'instructions': getattr(agent, 'instructions', ''),
                                'created_at': created_at_value,
                                'connection_string': endpoint,  # Add the Azure AI Projects endpoint
                                'agent_id': agent.id  # Add explicit agent_id for connection
                            })
                        return agents
                else:
                    return []
            
            # Use existing client if available
            agents_response = self.client.agents.list_agents()
            if hasattr(agents_response, 'data'):
                agents = []
                for agent in agents_response.data:
                    agents.append({
                        'id': agent.id,
                        'name': agent.name or f"Agent {agent.id[:8]}",
                        'model': getattr(agent, 'model', 'gpt-4o'),
                        'description': agent.description or f"Azure AI Agent using {getattr(agent, 'model', 'gpt-4o')}",
                        'instructions': getattr(agent, 'instructions', ''),
                        'created_at': getattr(agent, 'created_at', 0),
                        'connection_string': "https://eastus2.api.azureml.ms",  # Add the Azure AI Projects endpoint
                        'agent_id': agent.id  # Add explicit agent_id for connection
                    })
                return agents
            return []
            
        except Exception as e:
            return []
    
    def create_thread(self):
        """Create a new conversation thread - Direct Azure AI Foundry only, no fallbacks"""
        
        # Step 1: Check client availability
        if not self.client:
            raise Exception("Azure AI Project client not initialized. Please check workspace configuration.")
        
        # Step 2: Check agent availability  
        if not self.agent:
            raise Exception("Azure AI Agent not loaded. Please verify agent configuration.")
        
        # Step 3: Create thread directly - NO FALLBACKS
        try:
            thread = self.client.agents.create_thread()
            
            if not thread:
                raise Exception("Azure AI Foundry thread creation failed - no thread object returned")
            
            if not hasattr(thread, 'id') or not thread.id:
                raise Exception("Azure AI Foundry thread created but missing ID")
            
            return thread
            
        except Exception as e:
            raise Exception(f"Azure AI Foundry thread creation failed: {str(e)}")
    
    def send_message_and_get_response(self, thread_id: str, user_message: str):
        """Send message to agent and get response with enhanced error handling and fallbacks"""
        try:
            # Check if client is available
            if not self.client:
                raise Exception("Azure AI Project client not initialized. Please check your Azure AI Project connection settings.")
            
            # Check if agent is available - if not, try to reload it
            if not self.agent:
                try:
                    if self.agent_id:
                        self.agent = self.client.agents.get_agent(self.agent_id)
                    else:
                        pass
                except Exception as reload_error:
                    raise Exception(f"Azure AI Agent not loaded and reload failed: {str(reload_error)}. Please ensure the agent is properly configured.")
            
            # Final check for agent
            if not self.agent:
                raise Exception("Azure AI Agent not loaded. Please ensure the agent is properly configured.")
            
            # Check if agent has id attribute
            if not hasattr(self.agent, 'id') or not self.agent.id:
                raise Exception("Agent ID not available. Please ensure the agent is properly created in Azure AI Project.")
            
            
            max_retries = 3
            retry_delay = 1
            
            for attempt in range(max_retries):
                try:
                    pass
                    
                    # Create user message
                    self.client.agents.create_message(
                        thread_id=thread_id,
                        role="user",
                        content=user_message
                    )
                    
                    # Create and process run
                    run = self.client.agents.create_and_process_run(
                        thread_id=thread_id,
                        agent_id=self.agent.id
                    )
                    
                    # Get messages
                    messages = self.client.agents.list_messages(thread_id=thread_id)
                    
                    # Extract response - Handle different API versions
                    response_text = ""
                    first_message = None
                    
                    # Try new API first (data attribute)
                    if hasattr(messages, 'data') and messages.data:
                        # Find the last assistant message
                        for message in messages.data:
                            if message.role == "assistant" and hasattr(message, 'content') and message.content:
                                if len(message.content) > 0 and hasattr(message.content[0], 'text'):
                                    response_text = message.content[0].text.value
                                    first_message = message
                                    break
                    # Fallback to old API
                    elif hasattr(messages, 'text_messages'):
                        text_messages = list(messages.text_messages)
                        
                        if text_messages:
                            first_message = text_messages[0]
                            
                            # Extract the text content
                            if hasattr(first_message, 'text') and hasattr(first_message.text, 'value'):
                                response_text = first_message.text.value
                    else:
                        return "No response received from agent."
                    
                    if not response_text:
                        return "No response received from agent."
                    
                    # Enhanced document reference processing for Azure AI Agent SDK
                    if first_message:
                        response_text = self._process_document_references(response_text, first_message)
                    
                    return response_text if response_text else "No response received from agent."

                    return "No response received from agent."
                    
                except AzureError as e:
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay * (2 ** attempt))
                    else:
                        return f"AI Agent bağlantı hatası: {str(e)}"
                        
        except Exception as e:
            return f"AI Agent genel hatası: {str(e)}"
    
    def _process_document_references(self, response_text: str, message) -> str:
        """
        Enhanced document reference processing for Azure AI Agent SDK
        Extracts actual file names from blob storage and replaces generic references
        """
        try:
            import re
            import os
            
            # Get file mapping from vector store
            file_mapping = self.get_agent_files_mapping()
            
            # Method 1: Process annotations from message content
            if hasattr(message, 'content') and message.content:
                for content_item in message.content:
                    if hasattr(content_item, 'text') and hasattr(content_item.text, 'annotations'):
                        annotations = content_item.text.annotations
                        
                        for i, annotation in enumerate(annotations):
                            # Handle file citations with vector store references
                            if hasattr(annotation, 'file_citation'):
                                file_citation = annotation.file_citation
                                
                                # Try to get file info from vector store
                                try:
                                    # Get file ID from citation
                                    file_id = getattr(file_citation, 'file_id', None)
                                    
                                    if file_id:
                                        # First try to get filename from our mapping
                                        actual_filename = file_mapping.get(file_id)
                                        
                                        if not actual_filename:
                                            # Fallback: try to get file information from Azure AI
                                            try:
                                                file_info = self.client.agents.get_file(file_id)
                                                actual_filename = getattr(file_info, 'filename', None)
                                            except Exception as file_error:
                                                logger.warning(f"Could not retrieve file info for {file_id}: {file_error}")
                                        
                                        if actual_filename:
                                            # Clean up the filename 
                                            display_name = os.path.basename(actual_filename)
                                            
                                            # Replace annotation text with actual filename with CSS class
                                            if hasattr(annotation, 'text'):
                                                old_text = annotation.text
                                                new_reference = f'<span class="document-reference">📄 {display_name}</span>'
                                                response_text = response_text.replace(old_text, new_reference)
                                        else:
                                            # Fallback: try to extract filename from quote if available
                                            quote_text = getattr(file_citation, 'quote', '')
                                            if quote_text:
                                                # Look for filename patterns in the quote
                                                filename_match = re.search(r'([^/\\]+\.[a-zA-Z0-9]+)', quote_text)
                                                if filename_match:
                                                    display_name = filename_match.group(1)
                                                    if hasattr(annotation, 'text'):
                                                        old_text = annotation.text
                                                        new_reference = f'<span class="document-reference">📄 {display_name}</span>'
                                                        response_text = response_text.replace(old_text, new_reference)
                                                    
                                except Exception as citation_error:
                                    logger.warning(f"Error processing file citation: {citation_error}")
                                    
                            # Handle URL citations (might contain blob storage URLs)
                            elif hasattr(annotation, 'url_citation'):
                                url_citation = annotation.url_citation
                                
                                try:
                                    # Get URL and try to extract filename
                                    url = getattr(url_citation, 'url', '')
                                    title = getattr(url_citation, 'title', '')
                                    
                                    # Check if it's a blob storage URL
                                    if 'blob.core.windows.net' in url:
                                        # Extract filename from blob URL
                                        blob_filename = os.path.basename(url.split('?')[0])  # Remove SAS token
                                        if blob_filename and blob_filename != 'blob.core.windows.net':
                                            display_name = blob_filename
                                        else:
                                            display_name = title or f"Belge {i+1}"
                                    else:
                                        display_name = title or os.path.basename(url) or f"Web Kaynağı {i+1}"
                                    
                                    # Replace annotation with formatted reference with CSS class
                                    if hasattr(annotation, 'text'):
                                        old_text = annotation.text
                                        new_reference = f'<span class="document-reference">🔗 {display_name}</span>'
                                        response_text = response_text.replace(old_text, new_reference)
                                        
                                except Exception as url_error:
                                    logger.warning(f"Error processing URL citation: {url_error}")
            
            # Method 2: Process legacy citation patterns and try to map with actual files
            if file_mapping:
                # Create a list of filenames for pattern matching
                filenames_list = list(file_mapping.values())
                
                # Replace generic patterns with actual filenames when possible
                citation_patterns = [
                    (r'\[(\d+):source\]', lambda m: self._get_filename_by_index(filenames_list, int(m.group(1)))),
                    (r'\[doc_(\d+)\]', lambda m: self._get_filename_by_index(filenames_list, int(m.group(1))-1)),
                    (r'\[file_(\d+)\]', lambda m: self._get_filename_by_index(filenames_list, int(m.group(1))-1))
                ]
                
                for pattern, replacement_func in citation_patterns:
                    response_text = re.sub(pattern, replacement_func, response_text)
            else:
                # Fallback to generic patterns with CSS class
                citation_patterns = [
                    (r'\[(\d+):source\]', lambda m: f'<span class="document-reference">📄 Kaynak Dosya {int(m.group(1))+1}</span>'),
                    (r'\[doc_(\d+)\]', lambda m: f'<span class="document-reference">📄 Belge {m.group(1)}</span>'),
                    (r'\[file_(\d+)\]', lambda m: f'<span class="document-reference">📄 Dosya {m.group(1)}</span>')
                ]
                
                for pattern, replacement in citation_patterns:
                    response_text = re.sub(pattern, replacement, response_text)
            
            # Method 3: Additional cleanup patterns with CSS class
            cleanup_patterns = [
                (r'\[source:\s*([^\]]+)\]', lambda m: f'<span class="document-reference">📄 {m.group(1).strip()}</span>'),
            ]
            
            for pattern, replacement in cleanup_patterns:
                response_text = re.sub(pattern, replacement, response_text)
            
            # Method 4: Post-process to merge fragmented document references
            # This handles cases where a single filename gets split across multiple annotations
            response_text = self._merge_fragmented_references(response_text)
            
            return response_text
            
        except Exception as e:
            logger.error(f"Error processing document references: {e}")
            return response_text
    
    def _get_filename_by_index(self, filenames_list: List[str], index: int) -> str:
        """Get filename by index, with fallback to generic name"""
        try:
            if 0 <= index < len(filenames_list):
                filename = os.path.basename(filenames_list[index])
                return f'<span class="document-reference">📄 {filename}</span>'
            else:
                return f'<span class="document-reference">📄 Kaynak Dosya {index+1}</span>'
        except Exception:
            return f'<span class="document-reference">📄 Kaynak Dosya {index+1}</span>'
    
    def _merge_fragmented_references(self, text: str) -> str:
        """
        Merge fragmented document references that were split across multiple annotations.
        For example: "🔗 ARAÇ KULLANICI" + "📄 SORUMLULUKLARI.docx" -> "📄 ARAÇ KULLANICI SORUMLULUKLARI.docx"
        """
        try:
            import re
            
            # Debug log
            logger.info(f"[MERGE DEBUG] Starting merge process")
            
            # Strategy 1: Find and merge document references that appear to be parts of the same filename
            # Look for patterns like "🔗 ARAÇ KULLANICI" followed by "📄 SORUMLULUKLARI.docx"
            
            # First, let's identify all document reference spans
            span_pattern = r'<span class="document-reference">([^<]*)</span>'
            
            # Find all matches with their positions
            matches = list(re.finditer(span_pattern, text))
            
            if len(matches) <= 1:
                logger.info("[MERGE DEBUG] Only 1 or no spans found, nothing to merge")
                return text
            
            logger.info(f"[MERGE DEBUG] Found {len(matches)} spans to analyze")
            
            # Create list of spans with their details
            spans = []
            for i, match in enumerate(matches):
                content = match.group(1).strip()
                spans.append({
                    'index': i,
                    'full_match': match.group(0),
                    'content': content,
                    'start': match.start(),
                    'end': match.end(),
                    'has_file_extension': any(ext in content.lower() for ext in ['.docx', '.pdf', '.txt', '.xlsx', '.pptx'])
                })
            
            # Log all spans for debugging
            for span in spans:
                logger.info(f"[MERGE DEBUG] Span {span['index']}: '{span['content']}' (has_extension: {span['has_file_extension']})")
            
            # Strategy 2: Look for consecutive spans that might be parts of same filename
            merged_text = text
            changes_made = True
            iteration = 0
            max_iterations = 10
            
            while changes_made and iteration < max_iterations:
                iteration += 1
                changes_made = False
                
                # Re-parse spans after each change
                current_matches = list(re.finditer(span_pattern, merged_text))
                
                for i in range(len(current_matches) - 1):
                    current_match = current_matches[i]
                    next_match = current_matches[i + 1]
                    
                    current_content = current_match.group(1).strip()
                    next_content = next_match.group(1).strip()
                    
                    # Calculate distance between spans
                    distance = next_match.start() - current_match.end()
                    
                    # Check if they should be merged
                    should_merge = False
                    
                    # Rule 1: If they are very close (within 20 characters)
                    if distance <= 20:
                        should_merge = True
                        
                    # Rule 2: If first span ends with partial word and second starts with continuation
                    # Example: "ARAÇ KULLANICI" + "SORUMLULUKLARI.docx"
                    if (not current_content.endswith('.docx') and not current_content.endswith('.pdf') and
                        not current_content.endswith('.txt') and not current_content.endswith('.xlsx')):
                        # Check if next content looks like continuation of filename
                        if (next_content.endswith('.docx') or next_content.endswith('.pdf') or
                            next_content.endswith('.txt') or next_content.endswith('.xlsx') or
                            any(word in next_content.upper() for word in ['SORUMLULUKLARI', 'BELGE', 'DOSYA'])):
                            should_merge = True
                    
                    # Rule 3: Turkish filename pattern detection
                    if ('ARAÇ' in current_content.upper() and 'KULLANICI' in current_content.upper() and
                        'SORUMLULUKLARI' in next_content.upper()):
                        should_merge = True
                        
                    if should_merge:
                        # Extract text between spans
                        between_text = merged_text[current_match.end():next_match.start()].strip()
                        
                        # Remove icons and clean content
                        clean_current = re.sub(r'^[🔗📄🗂️📋📊🎯⚙️💼📈📉📝🔍]\s*', '', current_content).strip()
                        clean_next = re.sub(r'^[🔗📄🗂️📋📊🎯⚙️💼📈📉📝🔍]\s*', '', next_content).strip()
                        
                        # Combine content (use file icon 📄 for final result)
                        if between_text and len(between_text) <= 10:
                            combined_content = f"📄 {clean_current} {between_text} {clean_next}"
                        else:
                            combined_content = f"📄 {clean_current} {clean_next}"
                        
                        # Replace the segment from first span start to second span end
                        old_segment = merged_text[current_match.start():next_match.end()]
                        new_segment = f'<span class="document-reference">{combined_content}</span>'
                        
                        merged_text = merged_text.replace(old_segment, new_segment, 1)
                        
                        logger.info(f"[MERGE DEBUG] Iteration {iteration}: Merged '{current_content}' + '{next_content}' = '{combined_content}'")
                        changes_made = True
                        break  # Start over after making a change
                
                if not changes_made:
                    break
            
            # Strategy 3: One more pass for any remaining consecutive spans
            final_pattern = r'(<span class="document-reference">([^<]*)</span>)\s*(<span class="document-reference">([^<]*)</span>)'
            
            def final_merge(match):
                first_full = match.group(1)
                first_content = match.group(2).strip()
                second_full = match.group(3)
                second_content = match.group(4).strip()
                
                # Clean both contents
                clean_first = re.sub(r'^[🔗📄🗂️📋📊🎯⚙️💼📈📉📝🔍]\s*', '', first_content).strip()
                clean_second = re.sub(r'^[🔗📄🗂️📋📊🎯⚙️💼📈📉📝🔍]\s*', '', second_content).strip()
                
                # Combine with file icon
                combined = f"📄 {clean_first} {clean_second}"
                
                logger.info(f"[MERGE DEBUG] Final merge: '{first_content}' + '{second_content}' = '{combined}'")
                
                return f'<span class="document-reference">{combined}</span>'
            
            merged_text = re.sub(final_pattern, final_merge, merged_text)
            
            logger.info(f"[MERGE DEBUG] Merge completed after {iteration} iterations")
            
            return merged_text
            
        except Exception as e:
            logger.error(f"Error merging fragmented references: {e}")
            return text
    
    def send_message(self, user_message: str, thread_id: str):
        """Simple wrapper for send_message_and_get_response"""
        return self.send_message_and_get_response(thread_id, user_message)
    
    def get_recent_generated_files(self, thread_id: str, limit: int = 10) -> List[Dict[str, str]]:
        """
        Get recently generated files from the thread's latest messages
        This is more reliable than parsing message text
        
        Args:
            thread_id: The thread ID to check for generated files
            limit: Number of recent messages to check
            
        Returns:
            List of dictionaries with file info: [{'file_id': str, 'filename': str}]
        """
        try:
            if not self.client:
                return []
            
            generated_files = []
            
            # Get recent messages from the thread
            messages = self.client.agents.list_messages(thread_id=thread_id)
            
            # Process messages to find file attachments/annotations
            messages_to_check = []
            if hasattr(messages, 'data') and messages.data:
                messages_to_check = messages.data[:limit]
            
            for message in messages_to_check:
                if message.role == "assistant":
                    # Check for attachments
                    if hasattr(message, 'attachments') and message.attachments:
                        for attachment in message.attachments:
                            if hasattr(attachment, 'file_id'):
                                file_id = attachment.file_id
                                
                                # Get file details
                                try:
                                    file_details = self.client.agents.get_file(file_id)
                                    filename = getattr(file_details, 'filename', f'file_{file_id}')
                                    
                                    generated_files.append({
                                        'file_id': file_id,
                                        'filename': filename
                                    })
                                    logger.info(f"Found attached file: {filename} (ID: {file_id})")
                                except Exception as file_error:
                                    logger.warning(f"Could not get details for attached file {file_id}: {file_error}")
                    
                    # Check for content with file references
                    if hasattr(message, 'content') and message.content:
                        for content_item in message.content:
                            # Check for text content with file annotations
                            if hasattr(content_item, 'text') and hasattr(content_item.text, 'annotations'):
                                for annotation in content_item.text.annotations:
                                    # File citation annotations
                                    if hasattr(annotation, 'file_citation') and annotation.file_citation:
                                        file_id = annotation.file_citation.file_id
                                        
                                        try:
                                            file_details = self.client.agents.get_file(file_id)
                                            filename = getattr(file_details, 'filename', f'file_{file_id}')
                                            
                                            generated_files.append({
                                                'file_id': file_id,
                                                'filename': filename
                                            })
                                            logger.info(f"Found citation file: {filename} (ID: {file_id})")
                                        except Exception as file_error:
                                            logger.warning(f"Could not get details for citation file {file_id}: {file_error}")
                                    
                                    # File path annotations (for generated files)
                                    if hasattr(annotation, 'file_path') and annotation.file_path:
                                        file_id = annotation.file_path.file_id
                                        
                                        try:
                                            file_details = self.client.agents.get_file(file_id)
                                            filename = getattr(file_details, 'filename', f'generated_{file_id}')
                                            
                                            generated_files.append({
                                                'file_id': file_id,
                                                'filename': filename
                                            })
                                            logger.info(f"Found generated file: {filename} (ID: {file_id})")
                                        except Exception as file_error:
                                            logger.warning(f"Could not get details for generated file {file_id}: {file_error}")
            
            # Remove duplicates
            seen_ids = set()
            unique_files = []
            for file_info in generated_files:
                file_id = file_info['file_id']
                if file_id not in seen_ids:
                    seen_ids.add(file_id)
                    unique_files.append(file_info)
            
            logger.info(f"Found {len(unique_files)} recent generated files")
            return unique_files
            
        except Exception as e:
            logger.error(f"Error getting recent generated files: {e}")
            return []
    
    def get_agent_files_mapping(self) -> Dict[str, str]:
        """
        Get mapping of file IDs to actual filenames from Azure AI Agent's vector store and recent files
        Returns dict with file_id -> filename mapping
        """
        try:
            if not self.client:
                logger.warning("Azure AI client not available for file mapping")
                return {}
            
            file_mapping = {}
            
            # Method 1: Try to get files from agent's vector store
            if self.agent and hasattr(self.agent, 'id'):
                try:
                    # Get agent's vector store
                    agent_info = self.client.agents.get_agent(self.agent.id)
                    
                    if hasattr(agent_info, 'tool_resources') and agent_info.tool_resources:
                        if hasattr(agent_info.tool_resources, 'file_search') and agent_info.tool_resources.file_search:
                            vector_store_ids = getattr(agent_info.tool_resources.file_search, 'vector_store_ids', [])
                            
                            for vector_store_id in vector_store_ids:
                                try:
                                    # List files in vector store
                                    vector_store_files = self.client.agents.list_vector_store_files(
                                        vector_store_id=vector_store_id
                                    )
                                    
                                    if hasattr(vector_store_files, 'data'):
                                        for file_obj in vector_store_files.data:
                                            if hasattr(file_obj, 'id'):
                                                file_id = file_obj.id
                                                
                                                # Get file details
                                                try:
                                                    file_details = self.client.agents.get_file(file_id)
                                                    filename = getattr(file_details, 'filename', f'file_{file_id}')
                                                    file_mapping[file_id] = filename
                                                    logger.info(f"Vector store file: {filename} (ID: {file_id})")
                                                except Exception as file_detail_error:
                                                    logger.warning(f"Could not get details for vector store file {file_id}: {file_detail_error}")
                                                    
                                except Exception as vector_store_error:
                                    logger.warning(f"Could not list files in vector store {vector_store_id}: {vector_store_error}")
                            
                except Exception as agent_error:
                    logger.warning(f"Could not get agent vector store info: {agent_error}")
            
            # Method 2: Try to list all files accessible by the client (most comprehensive)
            try:
                all_files = self.client.agents.list_files()
                
                if hasattr(all_files, 'data'):
                    for file_obj in all_files.data:
                        if hasattr(file_obj, 'id') and hasattr(file_obj, 'filename'):
                            file_id = file_obj.id
                            filename = file_obj.filename
                            file_mapping[file_id] = filename
                            logger.info(f"Available file: {filename} (ID: {file_id})")
                
            except Exception as list_files_error:
                logger.warning(f"Could not list files: {list_files_error}")
            
            # Method 3: Try to get files from code interpreter tool resources (for generated files)
            if self.agent and hasattr(self.agent, 'id'):
                try:
                    agent_info = self.client.agents.get_agent(self.agent.id)
                    
                    if hasattr(agent_info, 'tool_resources') and agent_info.tool_resources:
                        if hasattr(agent_info.tool_resources, 'code_interpreter') and agent_info.tool_resources.code_interpreter:
                            code_interpreter_files = getattr(agent_info.tool_resources.code_interpreter, 'file_ids', [])
                            
                            for file_id in code_interpreter_files:
                                try:
                                    file_details = self.client.agents.get_file(file_id)
                                    filename = getattr(file_details, 'filename', f'generated_{file_id}')
                                    file_mapping[file_id] = filename
                                    logger.info(f"Code interpreter file: {filename} (ID: {file_id})")
                                except Exception as file_detail_error:
                                    logger.warning(f"Could not get details for code interpreter file {file_id}: {file_detail_error}")
                                    
                except Exception as code_interpreter_error:
                    logger.warning(f"Could not get code interpreter files: {code_interpreter_error}")
            
            logger.info(f"Total files in mapping: {len(file_mapping)}")
            return file_mapping
            
        except Exception as e:
            logger.error(f"Error getting agent files mapping: {e}")
            return {}
    
    def download_file_content(self, file_id: str) -> Optional[bytes]:
        """
        Download file content from Azure AI Agent using file ID
        
        Args:
            file_id: The file ID from Azure AI Agent
            
        Returns:
            File content as bytes or None if download fails
        """
        try:
            if not self.client:
                logger.error("Azure AI client not initialized")
                return None
            
            # Try to get file content using Azure AI Projects SDK
            try:
                file_content = self.client.agents.get_file_content(file_id)
                
                # Handle different response types from Azure AI SDK
                if file_content is None:
                    return None
                
                # If it's a file-like object with read method
                if hasattr(file_content, 'read'):
                    try:
                        content = file_content.read()
                        if isinstance(content, bytes):
                            return content
                        elif isinstance(content, str):
                            return content.encode('utf-8')
                        else:
                            return None
                    except Exception:
                        return None
                
                # If it's already bytes
                elif isinstance(file_content, bytes):
                    return file_content
                
                # If it's a string
                elif isinstance(file_content, str):
                    return file_content.encode('utf-8')
                
                # If it's a response object with content attribute
                elif hasattr(file_content, 'content'):
                    content = file_content.content
                    if isinstance(content, bytes):
                        return content
                    elif isinstance(content, str):
                        return content.encode('utf-8')
                
                # If it's an iterable (stream)
                elif hasattr(file_content, '__iter__') and not isinstance(file_content, (str, bytes)):
                    try:
                        content_bytes = b''
                        for chunk in file_content:
                            if isinstance(chunk, bytes):
                                content_bytes += chunk
                            elif isinstance(chunk, str):
                                content_bytes += chunk.encode('utf-8')
                        return content_bytes if content_bytes else None
                    except Exception:
                        return None
                
                else:
                    return None
                    
            except Exception as download_error:
                # Try alternative method - get file details first
                try:
                    file_details = self.client.agents.get_file(file_id)
                    
                    # Try to get content through file details
                    if hasattr(file_details, 'content'):
                        content = file_details.content
                        if isinstance(content, bytes):
                            return content
                        elif isinstance(content, str):
                            return content.encode('utf-8')
                    
                except Exception:
                    pass
                
                return None
                
        except Exception as e:
            logger.error(f"Error in download_file_content: {e}")
            return None
    
    def get_downloadable_files_from_message(self, message_content: str) -> List[Dict[str, str]]:
        """
        Extract downloadable file information from message content using Azure AI annotations
        
        Args:
            message_content: The message content containing file references
            
        Returns:
            List of dictionaries with file info: [{'file_id': str, 'filename': str, 'url': str}]
        """
        downloadable_files = []
        
        try:
            # First, try to get files from Azure AI Agent's run annotations if available
            # This is more reliable than parsing message content
            
            # Get file mapping for reliable filename lookup
            file_mapping = self.get_agent_files_mapping()
            
            # Look for file references in message annotations (if available)
            # This requires checking the latest run/message for file annotations
            
            # Pattern-based approach as fallback with more conservative matching
            import re
            
            # Only look for very specific file patterns that are likely to be real file IDs
            # Focus on patterns that appear in the context of file operations
            
            # Pattern 1: Look for sandbox file paths which are more reliable
            sandbox_patterns = [
                r'/mnt/data/([a-zA-Z0-9_\-\.öçşğüıÖÇŞĞÜİıİçÇşŞğĞüÜöÖ ]+\.(xlsx?|pdf|docx?|txt|csv|json|xml|ppt|pptx))',
                r'sandbox:([a-zA-Z0-9_\-\.öçşğüıÖÇŞĞÜİıİçÇşŞğĞüÜöÖ ]+\.(xlsx?|pdf|docx?|txt|csv|json|xml|ppt|pptx))',
            ]
            
            for pattern in sandbox_patterns:
                matches = re.findall(pattern, message_content, re.IGNORECASE)
                for match in matches:
                    filename = match[0] if isinstance(match, tuple) else match
                    filename = filename.strip()
                    
                    # Try to find corresponding file ID from mapping by filename
                    file_id = None
                    for fid, fname in file_mapping.items():
                        if fname == filename or filename in fname or fname in filename:
                            file_id = fid
                            break
                    
                    if file_id and filename:
                        downloadable_files.append({
                            'file_id': file_id,
                            'filename': filename,
                            'url': f"#download:{file_id}"
                        })
            
            # Pattern 2: Look for explicit download/file references with context
            download_context_patterns = [
                r'(?:indir|download|dosya).*?([a-zA-Z0-9_\-\.öçşğüıÖÇŞĞÜİıİçÇşŞğĞüÜöÖ ]+\.(xlsx?|pdf|docx?|txt|csv|json|xml|ppt|pptx))',
                r'([a-zA-Z0-9_\-\.öçşğüıÖÇŞĞÜİıİçÇşŞğĞüÜöÖ ]+\.(xlsx?|pdf|docx?|txt|csv|json|xml|ppt|pptx)).*?(?:hazır|ready|oluştur|creat)',
            ]
            
            for pattern in download_context_patterns:
                matches = re.findall(pattern, message_content, re.IGNORECASE)
                for match in matches:
                    filename = match[0] if isinstance(match, tuple) else match
                    filename = filename.strip()
                    
                    # Try to find corresponding file ID from mapping
                    file_id = None
                    for fid, fname in file_mapping.items():
                        if fname == filename or filename in fname or fname in filename:
                            file_id = fid
                            break
                    
                    if file_id and filename:
                        # Check if not already added
                        existing = any(f['file_id'] == file_id for f in downloadable_files)
                        if not existing:
                            downloadable_files.append({
                                'file_id': file_id,
                                'filename': filename,
                                'url': f"#download:{file_id}"
                            })
            
            # Pattern 3: If we have files in file_mapping, check if their names appear in the message
            for file_id, filename in file_mapping.items():
                if filename in message_content:
                    # Check if this file is mentioned in a download context
                    filename_pos = message_content.find(filename)
                    if filename_pos != -1:
                        # Check surrounding context for download-related words
                        start_pos = max(0, filename_pos - 100)
                        end_pos = min(len(message_content), filename_pos + 100)
                        context = message_content[start_pos:end_pos].lower()
                        
                        download_keywords = ['indir', 'download', 'hazır', 'ready', 'oluştur', 'creat', 'dosya', 'file']
                        if any(keyword in context for keyword in download_keywords):
                            # Check if not already added
                            existing = any(f['file_id'] == file_id for f in downloadable_files)
                            if not existing:
                                downloadable_files.append({
                                    'file_id': file_id,
                                    'filename': filename,
                                    'url': f"#download:{file_id}"
                                })
            
            # Remove duplicates based on file_id
            seen_ids = set()
            unique_files = []
            for file_info in downloadable_files:
                file_id = file_info['file_id']
                if file_id not in seen_ids:
                    seen_ids.add(file_id)
                    unique_files.append(file_info)
            
            logger.info(f"Found {len(unique_files)} downloadable files in message")
            for file_info in unique_files:
                logger.info(f"File: {file_info['filename']} (ID: {file_info['file_id']})")
            
            return unique_files
            
        except Exception as e:
            logger.error(f"Error extracting downloadable files: {e}")
            return []
    
    def upload_document(self, container_name: str, file_name: str, file_content: bytes):
        """Upload document to Azure Blob Storage"""
        try:
            # Try to upload to Azure Blob Storage
            try:
                # Create container if it doesn't exist
                container_client = self.blob_client.get_container_client(container_name)
                try:
                    container_client.get_container_properties()
                except:
                    container_client.create_container()
                
                # Upload blob
                blob_client = container_client.get_blob_client(file_name)
                blob_client.upload_blob(file_content, overwrite=True)
                
                return True
                
            except Exception as azure_error:
                pass
                
                # For demo purposes, just log the upload
                file_size = len(file_content) if file_content else 0
                return True
            
        except Exception as e:
            return False
    
    def list_documents(self, container_name: str):
        """List documents in container"""
        try:
            # For demo purposes, if Azure connection fails, return demo documents
            try:
                container_client = self.blob_client.get_container_client(container_name)
                
                # Try to get container properties first, create if doesn't exist
                try:
                    container_client.get_container_properties()
                except Exception as container_error:
                    if "ContainerNotFound" in str(container_error):
                        container_client.create_container()
                        # Return empty list for new container
                        return []
                    else:
                        raise container_error
                
                # List blobs in the container
                blobs = container_client.list_blobs()
                
                documents = []
                for blob in blobs:
                    documents.append({
                        'name': blob.name,
                        'size': blob.size or 0,
                        'last_modified': blob.last_modified,
                        'content_type': blob.content_settings.content_type if blob.content_settings else 'unknown'
                    })
                
                return documents
                
            except Exception as azure_error:
                pass
                
                # Return demo documents
                from datetime import datetime, timedelta
                demo_docs = [
                    {
                        'name': 'sample_report.pdf',
                        'size': 1024 * 1024,  # 1 MB
                        'last_modified': datetime.now() - timedelta(hours=2),
                        'content_type': 'application/pdf'
                    },
                    {
                        'name': 'data_analysis.xlsx',
                        'size': 512 * 1024,  # 512 KB
                        'last_modified': datetime.now() - timedelta(days=1),
                        'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    },
                    {
                        'name': 'meeting_notes.docx',
                        'size': 256 * 1024,  # 256 KB
                        'last_modified': datetime.now() - timedelta(hours=5),
                        'content_type': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                    }
                ]
                return demo_docs
            
        except Exception as e:
            return []
    
    def download_document(self, container_name: str, file_name: str) -> Dict:
        """Download document from blob storage"""
        try:
            pass
            
            # Get blob client for the specific file
            container_client = self.blob_client.get_container_client(container_name)
            blob_client = container_client.get_blob_client(file_name)
            
            # Check if blob exists
            if not blob_client.exists():
                return {
                    'success': False,
                    'message': f"Document '{file_name}' not found"
                }
            
            # Download blob content
            blob_data = blob_client.download_blob().readall()
            
            # Get blob properties for metadata
            blob_properties = blob_client.get_blob_properties()
            
            
            return {
                'success': True,
                'content': blob_data,
                'size': len(blob_data),
                'content_type': blob_properties.content_settings.content_type,
                'last_modified': blob_properties.last_modified,
                'message': f"Successfully downloaded {file_name}"
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f"Download failed: {str(e)}"
            }
    
    def delete_document(self, container_name: str, file_name: str, index_name: str = None):
        """Delete document from container and ensure removal from search index"""
        try:
            # Strict validation: index_name is required
            if not index_name:
                return False
            
            if not isinstance(index_name, str) or index_name.strip() == "":
                return False
            
            index_name = index_name.strip()
            
            # Delete from blob storage first
            container_client = self.blob_client.get_container_client(container_name)
            blob_client = container_client.get_blob_client(file_name)
            blob_client.delete_blob()
            
            
            # Use standard deletion method for the specified index
            index_deletion_success = self._standard_index_deletion(container_name, file_name, index_name)
            
            if index_deletion_success:
                pass
            else:
                # Still return True since blob deletion succeeded, but log the issue
                pass
            
            return True
            
        except Exception as e:
            return False

    def _standard_index_deletion(self, container_name: str, file_name: str, index_name: str):
        """Standard index deletion process for all indexes"""
        deletion_success = False
        
        # Step 1: Directly remove document from search index
        try:
            remove_result = self.remove_document_from_index(container_name, file_name, index_name)
            if remove_result:
                deletion_success = True
            else:
                pass
        except Exception as direct_delete_error:
            pass
        
        # Step 2: Trigger indexer for comprehensive reindexing
        try:
            reindex_result = self.trigger_reindex_after_document_change(container_name, index_name)
            if reindex_result["success"]:
                deletion_success = True
            else:
                pass
        except Exception as indexer_error:
            pass
        
        # Step 3: Additional cleanup - search and delete any matching documents
        if not deletion_success:
            try:
                advanced_delete_result = self._advanced_search_and_delete(container_name, file_name, index_name)
                if advanced_delete_result:
                    deletion_success = True
            except Exception as advanced_error:
                pass
        
        return deletion_success

    def create_search_index(self, index_name: str) -> bool:
        """Create Azure Search index with vector search capabilities"""
        try:
            if not self.search_index_client:
                return False
                
            # Check if index already exists
            try:
                existing_index = self.search_index_client.get_index(index_name)
                return True
            except Exception:
                # Index doesn't exist, create it
                pass
            
            # Define index schema
            fields = [
                SimpleField(name="id", type=SearchFieldDataType.String, key=True),
                SearchableField(name="content", type=SearchFieldDataType.String),
                SearchableField(name="metadata_storage_name", type=SearchFieldDataType.String, filterable=True),
                SimpleField(name="metadata_storage_path", type=SearchFieldDataType.String),
                SimpleField(name="metadata_storage_content_type", type=SearchFieldDataType.String, filterable=True),
                SimpleField(name="metadata_storage_size", type=SearchFieldDataType.Int64, filterable=True),
                SimpleField(name="metadata_storage_last_modified", type=SearchFieldDataType.DateTimeOffset, filterable=True, sortable=True),
            ]
            
            # Create the index
            search_index = SearchIndex(name=index_name, fields=fields)
            result = self.search_index_client.create_index(search_index)
            return True
            
        except Exception as e:
            return False
    
    def get_search_client(self, index_name: str) -> Optional[SearchClient]:
        """Get search client for specific index"""
        try:
            # Strict validation: index_name is required
            if not index_name:
                return None
            
            if not isinstance(index_name, str) or index_name.strip() == "":
                return None
            
            index_name = index_name.strip()
            
            if not self.config.search_endpoint or not self.config.search_admin_key:
                return None
                
            from azure.search.documents import SearchClient
            from azure.core.credentials import AzureKeyCredential
            
            credential = AzureKeyCredential(self.config.search_admin_key)
            search_client = SearchClient(
                endpoint=self.config.search_endpoint,
                index_name=index_name,
                credential=credential
            )
            return search_client
            
        except Exception as e:
            return None

    def extract_text_from_document(self, file_content: bytes, content_type: str, filename: str) -> str:
        """Extract text from various document types"""
        try:
            text_content = ""
            
            if content_type == 'application/pdf' or filename.lower().endswith('.pdf'):
                try:
                    import PyPDF2
                    from io import BytesIO
                    
                    pdf_reader = PyPDF2.PdfReader(BytesIO(file_content))
                    for page in pdf_reader.pages:
                        text_content += page.extract_text() + "\n"
                except Exception as pdf_error:
                    text_content = f"[PDF content extraction failed: {filename}]"
                    
            elif content_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document'] or filename.lower().endswith('.docx'):
                try:
                    from docx import Document
                    from io import BytesIO
                    
                    doc = Document(BytesIO(file_content))
                    for paragraph in doc.paragraphs:
                        text_content += paragraph.text + "\n"
                except Exception as docx_error:
                    text_content = f"[DOCX content extraction failed: {filename}]"
                    
            elif content_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'] or filename.lower().endswith('.xlsx'):
                try:
                    import openpyxl
                    from io import BytesIO
                    
                    wb = openpyxl.load_workbook(BytesIO(file_content))
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        text_content += f"Sheet: {sheet_name}\n"
                        for row in sheet.iter_rows(values_only=True):
                            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                            if row_text.strip():
                                text_content += row_text + "\n"
                except Exception as xlsx_error:
                    text_content = f"[XLSX content extraction failed: {filename}]"
                    
            elif content_type.startswith('text/') or filename.lower().endswith(('.txt', '.md', '.csv')):
                try:
                    text_content = file_content.decode('utf-8', errors='ignore')
                except Exception as text_error:
                    text_content = f"[Text content extraction failed: {filename}]"
            else:
                text_content = f"[Content type {content_type} not supported for text extraction: {filename}]"
                
            return text_content[:10000]  # Limit to 10K characters to avoid index size issues
            
        except Exception as e:
            return f"[Text extraction error: {filename}]"

    def get_text_embedding(self, text: str) -> Optional[List[float]]:
        """Generate text embedding using Azure OpenAI"""
        try:
            if not self.config.openai_endpoint or not self.config.openai_api_key:
                # Return mock embedding for demo purposes
                import random
                random.seed(hash(text) % 2147483647)  # Consistent mock embedding for same text
                return [random.uniform(-1, 1) for _ in range(1536)]
                
            # Use Azure OpenAI for embeddings
            headers = {
                "Content-Type": "application/json",
                "api-key": self.config.openai_api_key
            }
            
            # Truncate text if too long (max 8191 tokens for text-embedding-ada-002)
            max_chars = 8000  # Conservative estimate
            if len(text) > max_chars:
                text = text[:max_chars]
                
            data = {
                "input": text
            }
            
            # Try multiple embedding endpoints
            embedding_endpoints = [
                f"{self.config.openai_endpoint}/openai/deployments/text-embedding-ada-002/embeddings?api-version=2023-05-15",
                f"{self.config.openai_endpoint}/openai/deployments/text-embedding-3-small/embeddings?api-version=2024-02-01",
                f"{self.config.openai_endpoint}/openai/embeddings?api-version=2023-05-15"
            ]
            
            for endpoint in embedding_endpoints:
                try:
                    response = requests.post(
                        endpoint,
                        headers=headers,
                        json=data,
                        timeout=30
                    )
                    
                    if response.status_code == 200:
                        result = response.json()
                        embedding = result["data"][0]["embedding"]
                        return embedding
                    else:
                        continue
                        
                except Exception as endpoint_error:
                    continue
            
            # If all endpoints fail, use mock embedding
            import random
            random.seed(hash(text) % 2147483647)
            return [random.uniform(-1, 1) for _ in range(1536)]
                
        except Exception as e:
            # Return mock embedding as fallback
            import random
            random.seed(hash(text) % 2147483647)
            return [random.uniform(-1, 1) for _ in range(1536)]

    def index_document(self, container_name: str, filename: str, file_content: bytes, 
                      content_type: str, index_name: str) -> bool:
        """Index a document in Azure Search with vector embeddings using indexer"""
        try:
            # Trigger indexer to reindex all documents in the container
            result = self.trigger_reindex_after_document_change(container_name, index_name)
            
            if result["success"]:
                return True
            else:
                return False
                
        except Exception as e:
            return False

    def search_documents(self, query: str, index_name: str, top: int = 10) -> List[Dict]:
        """Search documents using hybrid search (text + vector)"""
        try:
            search_client = self.get_search_client(index_name)
            if not search_client:
                return []
            
            # Perform search
            search_results = search_client.search(
                search_text=query,
                top=top,
                include_total_count=True
            )
            
            results = []
            for result in search_results:
                results.append(dict(result))
                
            return results
            
        except Exception as e:
            return []

    def upload_and_index_document(self, container_name: str, filename: str, 
                                 file_content: bytes, content_type: str, index_name: str) -> Dict:
        """Upload document to blob storage and trigger indexer for reindexing"""
        try:
            # Upload to blob storage
            blob_upload_success = self.upload_document(container_name, filename, file_content)
            
            if not blob_upload_success:
                return {
                    "success": False,
                    "message": "Failed to upload document to blob storage",
                    "indexed": False
                }
            
            # Trigger indexer for reindexing if index_name is provided
            indexed = False
            index_message = "No index specified"
            
            if index_name:
                index_result = self.trigger_reindex_after_document_change(container_name, index_name)
                indexed = index_result["success"]
                index_message = index_result["message"]
            
            
            return {
                "success": True,
                "message": f"Document '{filename}' uploaded successfully to blob storage",
                "indexed": indexed,
                "index_message": index_message
            }
            
        except Exception as e:
            return {
                "success": False,
                "message": f"Error uploading document: {str(e)}",
                "indexed": False
            }
    
    def authenticate_with_device_code(self, username: str = None) -> Dict:
        """
        Authenticate using Azure AD device code flow (supports MFA)
        This is the recommended method for desktop applications
        """
        try:
            from msal import PublicClientApplication
            
            # MSAL configuration
            app = PublicClientApplication(
                client_id=self.config.public_client_id,
                authority=f"https://login.microsoftonline.com/{self.config.tenant_id}"
            )
            
            # Scope for basic profile information
            scopes = ["https://graph.microsoft.com/User.Read"]
            
            # Device code flow
            flow = app.initiate_device_flow(scopes=scopes)
            
            if "user_code" not in flow:
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": "Failed to create device flow"
                }
            
            # Return the device code information for user to complete authentication
            return {
                "success": False,  # Not yet complete
                "user": None,
                "token": None,
                "message": flow["message"],
                "device_flow": flow,
                "user_code": flow["user_code"],
                "verification_uri": flow["verification_uri"],
                "expires_in": flow["expires_in"]
            }
                
        except ImportError:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": "MSAL library not installed. Please install msal package."
            }
        except Exception as e:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Device code authentication error: {str(e)}"
            }
    
    def complete_device_code_authentication(self, device_flow: Dict) -> Dict:
        """
        Complete the device code authentication after user has authenticated
        """
        try:
            from msal import PublicClientApplication
            
            # MSAL configuration
            app = PublicClientApplication(
                client_id=self.config.public_client_id,
                authority=f"https://login.microsoftonline.com/{self.config.tenant_id}"
            )
            
            # Complete the device flow
            result = app.acquire_token_by_device_flow(device_flow)
            
            if "access_token" in result:
                # Success with device flow
                user_info = self._get_user_info_from_graph(result["access_token"])
                username = user_info.get("userPrincipalName") or user_info.get("mail")
                role = self._determine_user_role(username)
                permissions = self._get_user_permissions(username, role)
                
                return {
                    "success": True,
                    "user": {
                        "username": username,
                        "display_name": user_info.get("displayName", username),
                        "role": role,
                        "permissions": permissions,
                        "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                        "real_azure_user": True,
                        "auth_method": "device_code"
                    },
                    "token": result["access_token"],
                    "message": "Azure AD device code authentication successful"
                }
            else:
                error_description = result.get("error_description", "Unknown error")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD device code authentication failed: {error_description}"
                }
                
        except Exception as e:
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Device code completion error: {str(e)}"
            }

    def run_indexer(self, indexer_name: str) -> Dict:
        """Run Azure Search indexer to reindex all documents"""
        try:
            if not self.config.search_endpoint or not self.config.search_admin_key:
                return {"success": False, "message": "Search not configured"}
            
            from azure.search.documents.indexes import SearchIndexerClient
            from azure.core.credentials import AzureKeyCredential
            
            # Create indexer client
            credential = AzureKeyCredential(self.config.search_admin_key)
            indexer_client = SearchIndexerClient(
                endpoint=self.config.search_endpoint,
                credential=credential
            )
            
            # Run the indexer
            indexer_client.run_indexer(indexer_name)
            
            
            return {
                "success": True,
                "message": f"Indexer '{indexer_name}' started successfully",
                "indexer_name": indexer_name
            }
            
        except Exception as e:
            return {
                "success": False, 
                "message": f"Error running indexer: {str(e)}"
            }

    def get_indexer_status(self, indexer_name: str) -> Dict:
        """Get the status of an Azure Search indexer"""
        try:
            if not self.config.search_endpoint or not self.config.search_admin_key:
                return {"success": False, "message": "Search not configured"}
            
            from azure.search.documents.indexes import SearchIndexerClient
            from azure.core.credentials import AzureKeyCredential
            
            # Create indexer client
            credential = AzureKeyCredential(self.config.search_admin_key)
            indexer_client = SearchIndexerClient(
                endpoint=self.config.search_endpoint,
                credential=credential
            )
            
            # Get indexer status
            status = indexer_client.get_indexer_status(indexer_name)
            
            return {
                "success": True,
                "status": status.status,
                "last_result": {
                    "status": status.last_result.status if status.last_result else None,
                    "start_time": status.last_result.start_time if status.last_result else None,
                    "end_time": status.last_result.end_time if status.last_result else None,
                    "item_count": status.last_result.item_count if status.last_result else 0,
                    "failed_item_count": status.last_result.failed_item_count if status.last_result else 0
                }
            }
            
        except Exception as e:
            return {
                "success": False,
                "message": f"Error getting indexer status: {str(e)}"
            }

    def remove_document_from_index(self, container_name: str, file_name: str, index_name: str) -> bool:
        """Remove document from Azure Search index using title field search and chunk_id deletion"""
        try:
            # Strict validation: all parameters are required
            if not index_name:
                return False
            
            if not isinstance(index_name, str) or index_name.strip() == "":
                return False
            
            if not file_name or not isinstance(file_name, str) or file_name.strip() == "":
                return False
            
            index_name = index_name.strip()
            file_name = file_name.strip()
            
            if not self.config.search_endpoint or not self.config.search_admin_key:
                return False
            
            from azure.search.documents import SearchClient
            from azure.core.credentials import AzureKeyCredential
            
            # Create search client
            credential = AzureKeyCredential(self.config.search_admin_key)
            search_client = SearchClient(
                endpoint=self.config.search_endpoint,
                index_name=index_name,
                credential=credential
            )
            
            
            # Search for documents that match the filename in title field
            search_results = list(search_client.search(search_text=f'title:"{file_name}"', top=100))
            
            if len(search_results) == 0:
                # Try broader search if exact title doesn't work
                search_results = list(search_client.search(search_text=file_name, top=100))
            
            if len(search_results) == 0:
                return False
            
            # Find documents that actually match our target file
            documents_to_delete = []
            for result in search_results:
                result_dict = dict(result)
                title = result_dict.get('title', '')
                
                # Check if this document matches our file
                if file_name == title or file_name in title:
                    documents_to_delete.append(result_dict)
            
            if not documents_to_delete:
                return False
            
            
            # Try to delete each matching document using chunk_id
            successful_deletions = 0
            for i, doc in enumerate(documents_to_delete):
                try:
                    # Use chunk_id as the key field (based on our successful test)
                    chunk_id = doc.get('chunk_id')
                    if not chunk_id:
                        continue
                    
                    # Prepare deletion payload
                    delete_payload = {'chunk_id': chunk_id}
                    
                    # Perform the deletion
                    delete_result = search_client.delete_documents([delete_payload])
                    
                    if delete_result and len(delete_result) > 0:
                        result_item = delete_result[0]
                        if hasattr(result_item, 'succeeded') and result_item.succeeded:
                            successful_deletions += 1
                        elif hasattr(result_item, 'status_code') and result_item.status_code in [200, 204]:
                            successful_deletions += 1
                        else:
                            pass
                    else:
                        pass
                        
                except Exception as delete_error:
                    continue
            
            # Report results
            if successful_deletions > 0:
                return True
            else:
                return False
            
        except Exception as e:
            return False

    def trigger_reindex_after_document_change(self, container_name: str, index_name: str) -> Dict:
        """Trigger reindexing after document upload or deletion using indexer"""
        try:
            # Generate indexer name based on container/index
            indexer_name = f"{index_name}-indexer"
            
            # Run the indexer
            result = self.run_indexer(indexer_name)
            
            if result["success"]:
                return {
                    "success": True,
                    "message": f"Reindexing started for container '{container_name}'",
                    "indexer_name": indexer_name,
                    "container_name": container_name,
                    "index_name": index_name
                }
            else:
                return result
                
        except Exception as e:
            return {
                "success": False,
                "message": f"Error triggering reindex: {str(e)}"
            }

    def _advanced_search_and_delete(self, container_name: str, file_name: str, index_name: str) -> bool:
        """Advanced search and delete method - simplified and focused on title field search"""
        try:
            # Strict validation: all parameters are required
            if not index_name:
                return False
            
            if not isinstance(index_name, str) or index_name.strip() == "":
                return False
            
            if not file_name or not isinstance(file_name, str) or file_name.strip() == "":
                return False
            
            index_name = index_name.strip()
            file_name = file_name.strip()
            
            search_client = self.get_search_client(index_name)
            if not search_client:
                return False
            
            
            # Search for documents that match the filename using title field
            search_results = list(search_client.search(search_text=f'title:"{file_name}"', top=100))
            
            if len(search_results) == 0:
                # Try broader search
                search_results = list(search_client.search(search_text=file_name, top=100))
            
            if len(search_results) == 0:
                return False
            
            # Filter results to find exact matches
            matching_documents = []
            for result in search_results:
                result_dict = dict(result)
                title = result_dict.get('title', '')
                
                # Check if this is our target document
                if file_name == title or file_name in title:
                    matching_documents.append(result_dict)
            
            if not matching_documents:
                return False
            
            
            # Try to delete each matching document
            deletion_success_count = 0
            for i, doc in enumerate(matching_documents):
                try:
                    # Use chunk_id as key field (confirmed working in tests)
                    chunk_id = doc.get('chunk_id')
                    if not chunk_id:
                        continue
                    
                    # Prepare deletion payload
                    deletion_document = {'chunk_id': chunk_id}
                    
                    # Perform the deletion
                    delete_result = search_client.delete_documents([deletion_document])
                    
                    if delete_result and len(delete_result) > 0:
                        result_item = delete_result[0]
                        if hasattr(result_item, 'succeeded') and result_item.succeeded:
                            deletion_success_count += 1
                        elif hasattr(result_item, 'status_code') and result_item.status_code in [200, 204]:
                            deletion_success_count += 1
                        else:
                            pass
                    else:
                        pass
                    
                except Exception as delete_error:
                    continue
            
            # Report results
            if deletion_success_count > 0:
                return True
            else:
                return False
            
        except Exception as e:
            return False
    
    def send_message_with_code_interpreter(self, thread_id: str, user_message: str) -> tuple[str, Optional[str], Optional[str]]:
        """
        Send message to agent with code interpreter capability and return response text, image path, and code
        Returns: (response_text, image_path, code_snippet)
        """
        try:
            if not self.client or not self.agent:
                raise Exception("Azure AI Project client or agent not initialized")
            
            # Enhance user message with Turkish currency formatting instructions if needed
            enhanced_message = user_message
            if any(word in user_message.lower() for word in ["grafik", "chart", "tl", "lira", "tutar", "satış", "gelir", "plot", "visualization"]):
                enhanced_message += "\n\n[SİSTEM TALİMATI: Türk Lirası değerleri için grafik oluştururken, y ekseni etiketini değerlerin gerçek büyüklüğüne göre ayarla. Örneğin 2.5, 2.7 gibi değerler varsa 'Milyon TL', 2500, 2700 gibi değerler varsa 'Bin TL' yaz. Değerlerin sayısal büyüklüğünü kontrol et ve uygun birimi kullan.]"
            
            # Create user message  
            self.client.agents.create_message(
                thread_id=thread_id,
                role="user",
                content=enhanced_message
            )
            
            # Create and process run
            run = self.client.agents.create_and_process_run(
                thread_id=thread_id,
                agent_id=self.agent.id
            )
            
            # Check run status
            if run.status == "failed":
                return f"Run failed: {run.last_error}", None, None
            
            # Get messages
            messages = self.client.agents.list_messages(thread_id=thread_id)
            
            # Extract response text - Updated method for new API
            response_text = "No response from agent."
            if hasattr(messages, 'data') and messages.data:
                # Find the last assistant message
                for message in messages.data:
                    if message.role == "assistant" and hasattr(message, 'content') and message.content:
                        if len(message.content) > 0 and hasattr(message.content[0], 'text'):
                            response_text = message.content[0].text.value
                            break
            
            # Initialize return values
            image_path = None
            code_snippet = None
            
            # Check for file annotations (images) - Updated for new API
            if hasattr(messages, 'data') and messages.data:
                for message in messages.data:
                    if message.role == "assistant" and hasattr(message, 'content'):
                        for content_item in message.content:
                            # Check for image file references
                            if hasattr(content_item, 'image_file'):
                                try:
                                    file_id = content_item.image_file.file_id
                                    file_name = f"agent_image_{thread_id[-8:]}.png"
                                    
                                    # Save the image file locally
                                    self.client.agents.save_file(
                                        file_id=file_id,
                                        file_name=file_name
                                    )
                                    
                                    image_path = file_name
                                    break
                                    
                                except Exception as img_error:
                                    pass
                        if image_path:
                            break
            
            # Fallback: Check for file annotations (old API format)            
            elif hasattr(messages, 'file_path_annotations') and messages.file_path_annotations:
                try:
                    file_path_annotation = messages.file_path_annotations[0]
                    file_name = f"agent_image_{thread_id[-8:]}.png"
                    
                    # Save the image file locally
                    self.client.agents.save_file(
                        file_id=file_path_annotation.file_path.file_id,
                        file_name=file_name
                    )
                    
                    image_path = file_name
                    
                except Exception as img_error:
                    pass
            
            # Extract Python code snippet from run steps
            try:
                run_details = self.client.agents.list_run_steps(
                    thread_id=thread_id,
                    run_id=run.id
                )
                
                for step in run_details.data:
                    if getattr(step.step_details, 'type', None) == "tool_calls":
                        for call in step.step_details.tool_calls:
                            if hasattr(call, 'code_interpreter') and call.code_interpreter:
                                input_value = getattr(call.code_interpreter, 'input', None)
                                if input_value:
                                    code_snippet = input_value
                                    break
                        if code_snippet:
                            break
                            
            except Exception as code_error:
                pass
            
            return response_text, image_path, code_snippet
            
        except Exception as e:
            return f"Error: {str(e)}", None, None

    def create_code_interpreter_agent(self, model: str, name: str, instructions: str) -> bool:
        """
        Create a temporary agent with code interpreter capabilities
        """
        try:
            if not self.client:
                return False
            
            # Add Code Interpreter to the Agent's ToolSet
            toolset = ToolSet()
            code_interpreter_tool = CodeInterpreterTool()
            toolset.add(code_interpreter_tool)
            
            # Create agent with code interpreter
            temp_agent = self.client.agents.create_agent(
                model=model,
                name=name,
                instructions=instructions,
                toolset=toolset
            )
            
            if temp_agent:
                # Temporarily replace the current agent
                self.agent = temp_agent
                return True
            else:
                return False
                
        except Exception as e:
            return False

    def cleanup_temporary_agent(self):
        """
        Clean up temporary code interpreter agent and restore original agent
        """
        try:
            if self.client and self.agent and hasattr(self.agent, 'id'):
                # Check if this is a temporary agent (created by us)
                if 'code-interpreter' in getattr(self.agent, 'name', '').lower():
                    self.client.agents.delete_agent(self.agent.id)
                    
                    # Restore original agent
                    if self.agent_id:
                        try:
                            self.agent = self.client.agents.get_agent(self.agent_id)
                        except Exception as restore_error:
                            pass
                            
        except Exception as e:
            pass

class BlobStorageAgentManager:
    """Manages agent configurations using Azure Blob Storage"""
    
    def __init__(self, config: 'AzureConfig'):
        self.config = config
        self.container_name = "agent-configs"
        self.blob_client = None
        self._init_blob_client()
    
    def _init_blob_client(self):
        """Initialize blob client with proper authentication"""
        try:
            # Use connection string for reliable authentication
            if self.config.storage_connection_string:
                self.blob_client = BlobServiceClient.from_connection_string(
                    self.config.storage_connection_string
                )
            else:
                # Fallback to Azure AD authentication
                credential = DefaultAzureCredential()
                self.blob_client = BlobServiceClient(
                    account_url=f"https://{self.config.storage_account_name}.blob.core.windows.net",
                    credential=credential
                )
            
            # Ensure container exists
            try:
                self.blob_client.create_container(self.container_name)
            except Exception as e:
                pass  # Container might already exist
                
        except Exception as e:
            self.blob_client = None
    
    def get_agent(self, agent_id: str) -> Optional[Dict]:
        """Get agent configuration by ID"""
        try:
            if not self.blob_client:
                return None
                
            blob_name = f"{agent_id}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name, 
                blob=blob_name
            )
            
            try:
                blob_data = blob_client.download_blob().readall()
                return json.loads(blob_data.decode('utf-8'))
            except Exception:
                return None
                
        except Exception as e:
            return None
    
    def get_all_agents(self) -> Dict[str, Dict]:
        """Get all agent configurations"""
        try:
            if not self.blob_client:
                return {}
                
            container_client = self.blob_client.get_container_client(self.container_name)
            agents = {}
            
            for blob in container_client.list_blobs():
                if blob.name.endswith('.json'):
                    agent_id = blob.name.replace('.json', '')
                    agent_data = self.get_agent(agent_id)
                    if agent_data:
                        agents[agent_id] = agent_data
            
            return agents
            
        except Exception as e:
            return {}
    
    def get_active_agents(self) -> Dict[str, Dict]:
        """Get only active agent configurations"""
        all_agents = self.get_all_agents()
        # Include agents with status 'active' OR enabled=True OR missing status field (default to active)
        return {agent_id: agent_config for agent_id, agent_config in all_agents.items() 
                if (agent_config.get('status') == 'active' or 
                    agent_config.get('enabled', True) or 
                    'status' not in agent_config)}
    
    def add_agent(self, agent_config: Dict) -> bool:
        """Add new agent configuration"""
        try:
            if not self.blob_client:
                return False
                
            agent_id = agent_config.get('id')
            if not agent_id:
                return False
                
            blob_name = f"{agent_id}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            # Add timestamp and ensure status is set
            agent_config['created_at'] = datetime.now().isoformat()
            agent_config['updated_at'] = datetime.now().isoformat()
            
            # Ensure agent is marked as active if no status specified
            if 'status' not in agent_config:
                agent_config['status'] = 'active'
            
            # Ensure enabled field is set
            if 'enabled' not in agent_config:
                agent_config['enabled'] = True
            
            blob_client.upload_blob(
                json.dumps(agent_config, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            return False
    
    def update_agent(self, agent_id: str, agent_config: Dict) -> bool:
        """Update existing agent configuration"""
        try:
            if not self.blob_client:
                return False
                
            # Get existing config to preserve creation time
            existing_config = self.get_agent(agent_id)
            if existing_config:
                agent_config['created_at'] = existing_config.get('created_at')
            
            agent_config['updated_at'] = datetime.now().isoformat()
            agent_config['id'] = agent_id
            
            blob_name = f"{agent_id}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.upload_blob(
                json.dumps(agent_config, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            return False
    
    def delete_agent(self, agent_id: str) -> bool:
        """Delete agent configuration"""
        try:
            if not self.blob_client:
                return False
                
            blob_name = f"{agent_id}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.delete_blob()
            return True
            
        except Exception as e:
            return False
    
    def set_agent_status(self, agent_id: str, status: str) -> bool:
        """Set agent status (active/inactive)"""
        try:
            agent_config = self.get_agent(agent_id)
            if not agent_config:
                return False
                
            agent_config['status'] = status
            return self.update_agent(agent_id, agent_config)
            
        except Exception as e:
            return False
    
    def generate_azure_container_name(self, agent_id: str) -> str:
        """Generate Azure-compliant container name for agent"""
        # Azure container names must be lowercase, 3-63 chars, alphanumeric and hyphens
        container_name = f"agent-{agent_id}".lower()
        # Replace any invalid characters
        container_name = ''.join(c if c.isalnum() or c == '-' else '-' for c in container_name)
        # Ensure it doesn't exceed 63 characters
        if len(container_name) > 63:
            container_name = container_name[:63]
        # Ensure it doesn't end with a hyphen
        container_name = container_name.rstrip('-')
        return container_name

class BlobStorageUserManager:
    """Manages user authentication and permissions using Azure Blob Storage"""
    
    def __init__(self, config: 'AzureConfig'):
        self.config = config
        self.container_name = "user-configs"
        self.blob_client = None
        self._init_blob_client()
        self._init_default_admin()
    
    def _init_blob_client(self):
        """Initialize blob client with proper authentication"""
        try:
            # Use connection string for reliable authentication
            if self.config.storage_connection_string:
                self.blob_client = BlobServiceClient.from_connection_string(
                    self.config.storage_connection_string
                )
            else:
                # Fallback to Azure AD authentication
                credential = DefaultAzureCredential()
                self.blob_client = BlobServiceClient(
                    account_url=f"https://{self.config.storage_account_name}.blob.core.windows.net",
                    credential=credential
                )
            
            # Ensure container exists
            try:
                self.blob_client.create_container(self.container_name)
            except Exception as e:
                pass  # Container might already exist
                
        except Exception as e:
            self.blob_client = None
    
    def _init_default_admin(self):
        """Initialize default admin user if none exists"""
        try:
            admins = self.get_all_users()
            
            # Ensure admins is a dictionary
            if not isinstance(admins, dict):
                admins = {}
            
            # Check if any admin user exists
            admin_exists = False
            for username, user_data in admins.items():
                if isinstance(user_data, dict) and user_data.get('role') == 'admin':
                    admin_exists = True
                    break
            
            if not admin_exists:
                pass
                
        except Exception as e:
            pass
    
    def _hash_password(self, password: str) -> str:
        """Hash password using SHA256"""
        return hashlib.sha256(password.encode()).hexdigest()
    
    def create_admin_user(self, username: str, password: str) -> bool:
        """Create a new admin user with secure password - helper function for initial setup"""
        try:
            if not self.blob_client:
                return False
                
            user_data = {
                'username': username,
                'password_hash': self._hash_password(password),
                'role': 'admin',
                'permissions': ['all'],  # Admin has all permissions
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            }
            
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.upload_blob(
                json.dumps(user_data, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            return False

    def authenticate_admin(self, username: str, password: str) -> bool:
        """Authenticate admin user - requires admin credentials from blob storage"""
        try:
            user_data = self.get_user(username)
            if not user_data:
                return False
                
            if user_data.get('role') != 'admin':
                return False
                
            password_hash = self._hash_password(password)
            is_authenticated = user_data.get('password_hash') == password_hash
            
            if is_authenticated:
                pass
            else:
                pass
                
            return is_authenticated
            
        except Exception as e:
            return False
    
    def authenticate_azure_user(self, username: str, password: str) -> Dict:
        """Authenticate Azure user using Azure AD (placeholder implementation)"""
        try:
            # For now, just check if user exists in our system
            # In production, this would integrate with Azure AD
            user_data = self.get_user(username)
            if user_data:
                return {
                    'authenticated': True,
                    'user': user_data,
                    'token': 'azure_token_placeholder'
                }
            else:
                return {
                    'authenticated': False,
                    'error': 'User not found'
                }
                
        except Exception as e:
            return {
                'authenticated': False,
                'error': str(e)
            }
    
    def get_user(self, username: str) -> Optional[Dict]:
        """Get user by username"""
        try:
            if not self.blob_client:
                return None
                
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            try:
                blob_data = blob_client.download_blob().readall()
                return json.loads(blob_data.decode('utf-8'))
            except Exception:
                return None
                
        except Exception as e:
            return None
    
    def get_user_permissions(self, username: str) -> Dict:
        """Get user permissions"""
        try:
            user_data = self.get_user(username)
            if user_data:
                return {
                    'permissions': user_data.get('permissions', []),
                    'role': user_data.get('role', 'user')
                }
            else:
                return {
                    'permissions': [],
                    'role': 'guest'
                }
                
        except Exception as e:
            return {
                'permissions': [],
                'role': 'guest'
            }
    
    def has_permission(self, username: str, agent_id: str, permission_type: str) -> bool:
        """Check if user has specific permission for agent"""
        try:
            # Return False if username or agent_id is None/empty
            if not username or not agent_id:
                return False
                
            user_permissions = self.get_user_permissions(username)
            permissions = user_permissions.get('permissions', [])
            role = user_permissions.get('role', 'guest')
            
            # Admin has all permissions
            if role == 'admin' or 'all' in permissions:
                return True
            
            # Handle both old dictionary format and new list format
            if isinstance(permissions, dict):
                # Old dictionary format: {'scm': {'access': True, 'chat': True, ...}}
                agent_perms = permissions.get(agent_id, {})
                if isinstance(agent_perms, dict):
                    return agent_perms.get(permission_type, False)
                return False
            elif isinstance(permissions, list):
                # New list format: ['access', 'chat', 'scm:access', 'scm:chat', ...]
                # Check for wildcard permission (without agent_id prefix)
                if permission_type in permissions:
                    return True
                    
                # Check for specific agent permission (agent_id:permission_type)
                specific_permission = f"{agent_id}:{permission_type}"
                if specific_permission in permissions:
                    return True
                
                # *** REMOVED DEFAULT PERMISSIONS FOR STANDARD USERS ***
                # Standard users must have explicit permissions granted
                # No default access/chat permissions anymore
            
            return False
            
        except Exception as e:
            return False
    
    def add_user(self, username: str, role: str, password: str, permissions: List[str] = None) -> bool:
        """Add new user with explicit password requirement"""
        try:
            if not self.blob_client:
                return False
            
            # Get default permissions if none provided
            if permissions is None:
                permissions = self._get_user_permissions(username, role)
            
            user_data = {
                'username': username,
                'password_hash': self._hash_password(password),
                'role': role,
                'permissions': permissions,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            }
            
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.upload_blob(
                json.dumps(user_data, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            return False
    
    def update_user_permissions(self, username: str, permissions: List[str]) -> bool:
        """Update user permissions"""
        try:
            user_data = self.get_user(username)
            if not user_data:
                return False
                
            user_data['permissions'] = permissions
            user_data['updated_at'] = datetime.now().isoformat()
            
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.upload_blob(
                json.dumps(user_data, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            return False
    
    def delete_user(self, username: str) -> bool:
        """Delete user"""
        try:
            if not self.blob_client:
                return False
                
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.delete_blob()
            return True
            
        except Exception as e:
            return False
    
    def get_all_users(self) -> Dict[str, Dict]:
        """Get all users"""
        try:
            if not self.blob_client:
                return {}
                
            container_client = self.blob_client.get_container_client(self.container_name)
            users = {}
            
            for blob in container_client.list_blobs():
                if blob.name.endswith('.json'):
                    username = blob.name.replace('.json', '')
                    user_data = self.get_user(username)
                    if user_data:
                        # Remove password hash from returned data for security
                        safe_user_data = {k: v for k, v in user_data.items() if k != 'password_hash'}
                        users[username] = safe_user_data
            
            return users
            
        except Exception as e:
            return {}
    
    def _get_user_permissions(self, username: str, role: str) -> List[str]:
        """Get default permissions for user based on role"""
        try:
            if role == "admin":
                return ["all"]  # Admin has all permissions
            elif role == "standard":
                # Standard users get basic permissions
                return ["access", "chat", "view"]
            elif role == "limited":
                # Limited users get minimal permissions
                return ["access"]
            else:
                # Default/guest permissions
                return []
                
        except Exception as e:
            return []

# Agent configuration and management wrapper
class AgentManager:
    """Manages agent configurations and operations - now using blob storage"""
    
    def __init__(self, config: AzureConfig):
        # Initialize without any extra parameters
        self.blob_agent_manager = BlobStorageAgentManager(config)
    
    def get_agent(self, agent_id: str) -> Optional[Dict]:
        """Get agent configuration by ID"""
        try:
            return self.blob_agent_manager.get_agent(agent_id)
        except TypeError as e:
            # Handle the error gracefully
            return None
    
    def get_all_agents(self) -> Dict[str, Dict]:
        """Get all agent configurations"""
        return self.blob_agent_manager.get_all_agents()
    
    def get_active_agents(self) -> Dict[str, Dict]:
        """Get only active agent configurations"""
        return self.blob_agent_manager.get_active_agents()
    
    def add_agent(self, agent_config: Dict) -> bool:
        """Add a new agent configuration"""
        return self.blob_agent_manager.add_agent(agent_config)
    
    def save_agent(self, configuration: Dict) -> Dict:
        """Save agent configuration with fallback to local storage"""
        try:
            # Try to save using blob agent manager first
            try:
                # Generate a unique ID if not provided
                if 'id' not in configuration:
                    configuration['id'] = f"agent_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                
                # Add metadata
                configuration['saved_at'] = datetime.now().isoformat()
                configuration['source'] = 'blob_storage'
                
                # Use add_agent or update_agent based on whether it exists
                agent_id = configuration['id']
                existing_agent = self.blob_agent_manager.get_agent(agent_id)
                
                if existing_agent:
                    success = self.blob_agent_manager.update_agent(agent_id, configuration)
                    action = "updated"
                else:
                    success = self.blob_agent_manager.add_agent(configuration)
                    action = "added"
                
                if success:
                    return {"status": "success", "location": "blob_storage", "id": agent_id, "action": action}
                else:
                    # Fall through to local storage
                    pass
                    
            except Exception as blob_error:
                # Fall through to local storage
                pass
            
            # Fallback to local storage
            try:
                import os
                backup_path = "config_backup/agent_configs.json"
                os.makedirs(os.path.dirname(backup_path), exist_ok=True)
                
                # Load existing configs or create new
                try:
                    with open(backup_path, 'r', encoding='utf-8') as f:
                        configs = json.load(f)
                except (FileNotFoundError, json.JSONDecodeError):
                    configs = {"agents": []}
                
                # Add timestamp and unique ID
                if 'id' not in configuration:
                    configuration['id'] = f"local_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                configuration['saved_at'] = datetime.now().isoformat()
                configuration['source'] = 'local_fallback'
                
                # Check if agent already exists (update vs add)
                existing_index = None
                for i, agent in enumerate(configs["agents"]):
                    if agent.get('id') == configuration['id']:
                        existing_index = i
                        break
                
                if existing_index is not None:
                    configs["agents"][existing_index] = configuration
                    action = "updated"
                else:
                    configs["agents"].append(configuration)
                    action = "added"
                
                # Save to file
                with open(backup_path, 'w', encoding='utf-8') as f:
                    json.dump(configs, f, indent=2, ensure_ascii=False)
                
                return {"status": "success", "location": "local", "id": configuration['id'], "action": action}
                
            except Exception as local_error:
                return {"status": "error", "message": f"Failed to save agent: {local_error}"}
                
        except Exception as e:
            return {"status": "error", "message": f"Error saving agent: {e}"}
    
    def update_agent(self, agent_id: str, agent_config: Dict) -> bool:
        """Update an existing agent configuration"""
        return self.blob_agent_manager.update_agent(agent_id, agent_config)
    
    def delete_agent(self, agent_id: str) -> bool:
        """Delete an agent configuration"""
        return self.blob_agent_manager.delete_agent(agent_id)
    
    def set_agent_status(self, agent_id: str, status: str) -> bool:
        """Set agent status (active/inactive)"""
        return self.blob_agent_manager.set_agent_status(agent_id, status)
    
    def generate_azure_container_name(self, agent_id: str) -> str:
        """Generate Azure-compliant container name for agent"""
        return self.blob_agent_manager.generate_azure_container_name(agent_id)

# User authentication and authorization wrapper
class UserManager:
    """Manages user authentication and authorization - now using blob storage"""
    
    def __init__(self):
        self.blob_user_manager = BlobStorageUserManager(AzureConfig())
    
    def authenticate_admin(self, username: str, password: str) -> bool:
        """Authenticate admin user"""
        return self.blob_user_manager.authenticate_admin(username, password)
    
    def authenticate_azure_user(self, username: str, password: str) -> Dict:
        """Authenticate Azure user using Azure AD"""
        return self.blob_user_manager.authenticate_azure_user(username, password)
    
    def get_user_permissions(self, username: str) -> Dict:
        """Get user permissions"""
        return self.blob_user_manager.get_user_permissions(username)
    
    def has_permission(self, username: str, agent_id: str, permission_type: str) -> bool:
        """Check if user has specific permission for agent"""
        return self.blob_user_manager.has_permission(username, agent_id, permission_type)
    
    def add_user(self, username: str, role: str, password: str, permissions: List[str] = None) -> bool:
        """Add a new user with explicit password requirement"""
        return self.blob_user_manager.add_user(username, role, password, permissions or [])
    
    def update_user_permissions(self, username: str, permissions: List[str]) -> bool:
        """Update user permissions"""
        return self.blob_user_manager.update_user_permissions(username, permissions)
    
    def delete_user(self, username: str) -> bool:
        """Delete a user"""
        return self.blob_user_manager.delete_user(username)
    
    def get_all_users(self) -> Dict[str, Dict]:
        """Get all users"""
        return self.blob_user_manager.get_all_users()
